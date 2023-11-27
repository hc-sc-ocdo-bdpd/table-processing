from table_metrics import test_tables, accuracy
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np

def compare_tables(file1, file2):
    '''
    Compares two Excel files and generates a new Excel file with comparison 
    results of the tables within each sheet. Input should be the file path.
    Results are generated in use_tools > excel_table_comparison.

    Note: This function is based off of the accuracy function from table_metrics,
    so the # of correct cells is calculated as strictly the # of cells that are
    in the second table that are equal to the cells in the first.

    It does not account for cells that are in the first but not in the second, or 
    cells that are in the second (outside of the range of the first table) but not in the first.
    '''
    
    # Import Excel file paths
    file_path_original = file1
    file_path_comparison = file2
    original_wb = openpyxl.load_workbook(file_path_original)
    comparison_wb = openpyxl.load_workbook(file_path_comparison)

    # Create a new workbook and worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Summary"
    # Add column titles to the "Summary" sheet
    sheet.append(['Sheet name', 'Accuracy', 'Overlap', 'String Similarity', 'Completeness', 'Purity', 'Precision', 'Recall'])

    # Initialize list to store accuracy values
    accuracy_list = []

    # Go through each sheet in the comparison workbook
    for sheet_name in comparison_wb.sheetnames:
        if sheet_name in original_wb.sheetnames:
            original_df = pd.read_excel(file_path_original, sheet_name=sheet_name, header=None, keep_default_na=False)
            comparison_df = pd.read_excel(file_path_comparison, sheet_name=sheet_name, header=None, keep_default_na=False)

            # Write the sheet name to the worksheet
            new_sheet = workbook.create_sheet(title=sheet_name)

            # Calculate comparison metrics using the test_tables function
            metrics_df = test_tables({sheet_name: [original_df,comparison_df]})

            # Write the second table to the worksheet
            rows = dataframe_to_rows(comparison_df, index=False, header=False)
            for row in rows:
                new_sheet.append(row)

            # Create a dataframe differences_df that will contain True where the two tables are equal and False where they are not
            differences_df = pd.DataFrame(np.full_like(original_df, False))

            # Transform tables to lists of lists
            vals1 = original_df.values.tolist()
            vals2 = comparison_df.values.tolist()

            # Loop through values
            for i in range(0, len(vals1)):
                row = vals1[i]
                for j in range(0, len(row)):
                    cell1 = row[j]
                    # Verify if cell1 is contained in cell2 (completely identified)
                    try:
                        cell2 = vals2[i][j]
                        if cell1 == cell2:
                            # If the cells are equal, set the corresponding cell in differences_df to True
                            differences_df.iat[i, j] = True
                    except IndexError:
                        pass
                    
            # Format the cells that are different from the original table
            for row in new_sheet.iter_rows(min_row=new_sheet.max_row-len(comparison_df)+1, min_col=1, max_col=len(comparison_df.columns)):
                for cell in row:
                    # Check if the cell is within the differences_df range
                    if cell.row-new_sheet.max_row+len(comparison_df)-1 < len(differences_df) and cell.column-1 < len(differences_df.columns):
                        # Check if the cell is different from the original table
                        if differences_df.iat[cell.row-new_sheet.max_row+len(comparison_df)-1, cell.column-1] == False:
                            # If the cell is different, highlight it with yellow
                            cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # Calculate the statistics
            num_cells = original_df.size
            num_true = differences_df.sum().sum()

            # Write the statistics to the worksheet
            new_sheet.append([])
            new_sheet.append(['Number of Correct Cells (Cells equal to original table)', num_true])
            new_sheet.append(['Number of Incorrect Cells (Cells not equal to original table)', num_cells - num_true])
            new_sheet.append(['Accuracy', accuracy(original_df, comparison_df)])
            new_sheet.append([])
            new_sheet.append(['Comparison metrics:'])

            # Write comparison metrics to the Excel file
            rows = dataframe_to_rows(metrics_df, index=False, header=True)
            for row in rows:
                new_sheet.append(row)

            # Write sheet name and comparison metrics to "Summary" sheet
            sheet.append([sheet_name, metrics_df['Accuracy'][0], metrics_df['Overlap'][0], metrics_df['String Similarity'][0], 
                metrics_df['Completeness'][0], metrics_df['Purity'][0], metrics_df['Precision'][0], metrics_df['Recall'][0]])

            # Append % correct to list
            accuracy_list.append(accuracy(original_df, comparison_df))
        else:
            print(f"{sheet_name} not found in {file1}")

    # Calculate statistics for % correct over all sheets
    mean_accuracy = np.mean(accuracy_list)
    median_accuracy = np.median(accuracy_list)
    min_accuracy = np.min(accuracy_list)
    max_accuracy = np.max(accuracy_list)
    range_accuracy = max_accuracy - min_accuracy

    # Write statistics for % correct over all sheets to "Summary" sheet
    sheet.append([])
    sheet.append(['Summary Statistics:'])
    sheet.append(['Mean (Accuracy)', round(mean_accuracy, 3)])
    sheet.append(['Median (Accuracy)', round(median_accuracy, 3)])
    sheet.append(['Min (Accuracy)', round(min_accuracy, 3)])
    sheet.append(['Max (Accuracy)', round(max_accuracy, 3)])
    sheet.append(['Range (Accuracy)', round(range_accuracy, 3)])

    # Save the workbook to a new file
    save_path = './use_tools/excel_table_comparison/table_comparison_results.xlsx'
    workbook.save(save_path)

# Get file names from user input
first_file = input("Enter the name of the first Excel file: ")
second_file = input("Enter the name of the second Excel file: ")

# Use the function to compare the files
compare_tables(first_file, second_file)